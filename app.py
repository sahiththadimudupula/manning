import math
import re
from io import BytesIO
from pathlib import Path
from typing import Optional, Tuple

import altair as alt
import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from schema_utils import align_and_validate_schemas

st.set_page_config(
    page_title="Manpower Recommendation Engine",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="collapsed",
)

PRIMARY_SPINNING_PATH = Path("Spinning.xlsx")
FALLBACK_SPINNING_PATHS = [
    PRIMARY_SPINNING_PATH,
    Path(__file__).resolve().parent / "Spinning.xlsx",
    Path("Spinning.xlsx"),
    Path("/mnt/data/Spinning.xlsx"),
]

PRIMARY_LOGO_PATH = Path("smartOps logo.png")
FALLBACK_LOGO_PATHS = [
    PRIMARY_LOGO_PATH,
    Path(__file__).resolve().parent / "smartOps logo.png",
    Path("smartOps logo.png"),
    Path("/mnt/data/smartOps logo.png"),
]

DISPLAY_COLUMNS = [
    "Location",
    "Business",
    "Section",
    "Sr_No",
    "Dept_Machine_Name",
    "Designation",
    "Machine_Count",
    "Workload",
    "Formulas",
    "BE_Scientific_Manpower",
    "Operator_Type",
    "Contractors",
    "Company_Associate",
    "BE_Final_Manpower",
    "N_shifts",
    "General_Shift",
    "Shift_A",
    "Shift_B",
    "Shift_C",
    "Reliever",
    "Remarks",
]

EDITABLE_COLUMNS = [
    "Formulas",
    "BE_Final_Manpower",
    "General_Shift",
    "Shift_A",
    "Shift_B",
    "Shift_C",
    "Reliever",
    "Remarks",
]

NUMERIC_COLUMNS = [
    "Sr_No",
    "Machine_Count",
    "BE_Scientific_Manpower",
    "Contractors",
    "Company_Associate",
    "BE_Final_Manpower",
    "N_shifts",
    "General_Shift",
    "Shift_A",
    "Shift_B",
    "Shift_C",
    "Reliever",
]

TEXT_COLUMNS = [
    "Location",
    "Business",
    "Section",
    "Dept_Machine_Name",
    "Designation",
    "Workload",
    "Formulas",
    "Operator_Type",
    "Remarks",
]

st.markdown(
    """
    <style>
        :root {
            --primary: #1e40af;
            --primary-2: #2563eb;
            --primary-3: #38bdf8;
            --bg-1: #f4f8ff;
            --bg-2: #edf4ff;
            --surface: rgba(255,255,255,0.78);
            --surface-strong: rgba(255,255,255,0.92);
            --text-1: #0f172a;
            --text-2: #334155;
            --text-3: #64748b;
            --border: rgba(148, 163, 184, 0.18);
            --shadow-xs: 0 4px 12px rgba(15, 23, 42, 0.04);
            --shadow-sm: 0 10px 28px rgba(15, 23, 42, 0.07);
            --shadow-md: 0 18px 42px rgba(15, 23, 42, 0.10);
            --shadow-lg: 0 22px 54px rgba(30, 64, 175, 0.12);
        }

        html, body, [class*="css"] {
            font-family: "Inter", "Segoe UI", sans-serif;
        }

        header,
        [data-testid="stHeader"],
        [data-testid="stToolbar"],
        .stAppToolbar,
        #MainMenu,
        footer,
        [data-testid="stDecoration"],
        [data-testid="stStatusWidget"] {
            display: none !important;
            height: 0 !important;
            visibility: hidden !important;
        }

        .stApp {
            background:
                radial-gradient(circle at top left, rgba(56,189,248,0.08), transparent 22%),
                radial-gradient(circle at top right, rgba(37,99,235,0.09), transparent 25%),
                linear-gradient(180deg, var(--bg-1) 0%, var(--bg-2) 100%);
            color: var(--text-1);
        }

        [data-testid="stAppViewContainer"],
        [data-testid="stAppViewContainer"] > section,
        [data-testid="stAppViewContainer"] > section > div,
        section.main,
        section.main > div {
            padding-top: 0 !important;
            margin-top: 0 !important;
        }

        .block-container,
        .main .block-container {
            max-width: 1540px;
            padding-top: 0 !important;
            padding-bottom: 1rem !important;
            margin-top: 0 !important;
        }

        /* hard cleanup for top ghost spacing */
        .main .block-container > div:first-child,
        .main .block-container > div:nth-child(2) {
            margin-top: 0 !important;
            padding-top: 0 !important;
            min-height: 0 !important;
        }

        .main .block-container > div:first-child > div,
        .main .block-container > div:nth-child(2) > div {
            margin-top: 0 !important;
            padding-top: 0 !important;
            min-height: 0 !important;
        }

        [data-testid="stSkeleton"],
        .stSkeleton,
        div[data-testid="stVerticalBlock"] > div:empty,
        div[data-testid="stVerticalBlockBorderWrapper"]:empty,
        div[data-testid="stMarkdownContainer"]:empty,
        .element-container:empty {
            display: none !important;
            height: 0 !important;
            min-height: 0 !important;
            margin: 0 !important;
            padding: 0 !important;
        }

        .element-container {
            margin-bottom: 0.25rem !important;
        }

        hr {
            display: none !important;
        }

        label[data-testid="stWidgetLabel"] p {
            color: var(--text-1) !important;
            font-weight: 700 !important;
            font-size: 0.90rem !important;
        }

        div[data-testid="stTabs"] {
            margin-top: 0.15rem;
        }

        div[data-testid="stTabs"] [role="tablist"] {
            gap: 0.35rem;
            background: rgba(255,255,255,0.60);
            border: 1px solid rgba(148,163,184,0.16);
            padding: 0.35rem;
            border-radius: 16px;
            backdrop-filter: blur(12px);
            box-shadow: var(--shadow-xs);
        }

        div[data-testid="stTabs"] button {
            font-weight: 700;
            font-size: 0.92rem;
            padding: 0.62rem 0.92rem;
            color: var(--text-2) !important;
            border-radius: 12px !important;
            transition: all 0.22s ease;
        }

        div[data-testid="stTabs"] button:hover {
            background: rgba(37, 99, 235, 0.08) !important;
            color: var(--primary-2) !important;
        }

        div[data-testid="stTabs"] button[aria-selected="true"] {
            background: linear-gradient(135deg, #1e40af 0%, #2563eb 100%) !important;
            color: #ffffff !important;
            box-shadow: 0 10px 24px rgba(37, 99, 235, 0.22);
        }

        div[data-testid="stTabs"] button p {
            color: inherit !important;
        }

        div[data-baseweb="select"] > div,
        .stTextInput > div > div > input {
            min-height: 44px !important;
            background: rgba(255,255,255,0.82) !important;
            border: 1px solid rgba(148,163,184,0.22) !important;
            border-radius: 14px !important;
            box-shadow: 0 8px 20px rgba(15, 23, 42, 0.04);
            color: var(--text-1) !important;
            backdrop-filter: blur(12px);
        }

        div[data-baseweb="select"] > div * {
            color: var(--text-1) !important;
            opacity: 1 !important;
        }

        div[data-baseweb="select"] input {
            color: var(--text-1) !important;
            -webkit-text-fill-color: var(--text-1) !important;
        }

        div[data-baseweb="select"] svg {
            color: var(--text-2) !important;
            fill: var(--text-2) !important;
        }

        ul[role="listbox"] {
            background: rgba(255,255,255,0.96) !important;
            border-radius: 16px !important;
            border: 1px solid rgba(148,163,184,0.16);
            backdrop-filter: blur(14px);
            box-shadow: var(--shadow-md) !important;
        }

        ul[role="listbox"] li {
            color: var(--text-1) !important;
            background: transparent !important;
        }

        ul[role="listbox"] li[aria-selected="true"] {
            background: rgba(37, 99, 235, 0.10) !important;
            color: var(--primary) !important;
        }

        div[data-baseweb="tag"] {
            background: linear-gradient(135deg, var(--primary) 0%, var(--primary-2) 100%) !important;
            border: none !important;
            border-radius: 999px !important;
            padding: 3px 10px !important;
            box-shadow: 0 8px 18px rgba(37, 99, 235, 0.18);
        }

        div[data-baseweb="tag"] span,
        div[data-baseweb="tag"] svg {
            color: #ffffff !important;
            fill: #ffffff !important;
            font-weight: 700 !important;
        }

        .stButton > button,
        .stDownloadButton > button {
            width: 100%;
            min-height: 46px;
            border-radius: 14px;
            border: 1px solid rgba(255,255,255,0.18);
            background: linear-gradient(135deg, #1e40af 0%, #2563eb 58%, #38bdf8 100%);
            color: #ffffff !important;
            font-weight: 700;
            font-size: 0.93rem;
            box-shadow: 0 14px 30px rgba(37,99,235,0.20);
            transition: transform 0.18s ease, box-shadow 0.18s ease, filter 0.18s ease;
        }

        .stButton > button:hover,
        .stDownloadButton > button:hover {
            transform: translateY(-1px);
            box-shadow: 0 18px 36px rgba(37,99,235,0.25);
            filter: saturate(1.05);
            color: #ffffff !important;
        }

        div[data-testid="stExpander"] details {
            border-radius: 18px;
            border: 1px solid rgba(37,99,235,0.14);
            background: linear-gradient(135deg, #1e40af 0%, #2563eb 72%, #38bdf8 100%);
            overflow: hidden;
            box-shadow: var(--shadow-sm);
        }

        div[data-testid="stExpander"] details summary p,
        div[data-testid="stExpander"] details summary span {
            color: #ffffff !important;
            font-weight: 700;
        }

        div[data-testid="stExpanderDetails"] {
            background: rgba(255,255,255,0.96);
            border-radius: 0 0 18px 18px;
            padding: 0.85rem 0.45rem 0.35rem 0.45rem;
            backdrop-filter: blur(8px);
        }

        div[data-testid="stExpanderDetails"] * {
            color: var(--text-1) !important;
        }

        div[data-testid="stDataEditor"],
        div[data-testid="stDataFrame"] {
            border-radius: 18px;
            overflow: hidden;
            border: 1px solid rgba(148,163,184,0.18);
            background: rgba(255,255,255,0.82);
            backdrop-filter: blur(14px);
            box-shadow: var(--shadow-sm);
        }

        

        /* removed thin blue line */
        .hero-shell::after {
            display: none;
        }

        .hero-badge {
            display: inline-block;
            background: rgba(37, 99, 235, 0.08);
            color: var(--primary) !important;
            border: 1px solid rgba(37, 99, 235, 0.14);
            border-radius: 999px;
            padding: 0.20rem 0.58rem;
            font-size: 0.85rem !important;
            font-weight: 800;
            margin-bottom: 0.20rem;
            letter-spacing: 0.18px;
            backdrop-filter: blur(8px);
        }

        .hero-gradient-title {
    font-size: 2.2rem;
    font-weight: 900;
    background: linear-gradient(90deg, #1e40af, #2563eb, #38bdf8);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
            }
        
        .hero-title {
            font-size: 1.4rem !important;
            font-weight: 900;
            line-height: 1.02;
            color: var(--text-1) !important;
            margin-bottom: 0.10rem;
        }

        .hero-subtitle {
            font-size:  0.95rem !important;
            line-height: 1.40;
            color: var(--text-2) !important;
            font-weight: 500;
            max-width: 820px;
        }

        .hero-logo-card {
            background: rgba(255,255,255,0.78);
            border-radius: 14px;
            padding: 0.18rem;
            border: 1px solid rgba(255,255,255,0.52);
            box-shadow: 0 8px 18px rgba(15, 23, 42, 0.05);
            text-align: center;
            backdrop-filter: blur(12px);
        }

        .premium-strip {
            background: transparent;
            padding: 0;
            margin-bottom: 0.7rem;
        }

        .metric-card {
            background: linear-gradient(180deg, rgba(255,255,255,0.86) 0%, rgba(255,255,255,0.76) 100%);
            border-radius: 20px;
            padding: 0.88rem 1rem 0.82rem 1rem;
            box-shadow: var(--shadow-sm);
            border: 1px solid rgba(255,255,255,0.45);
            min-height: 108px;
            position: relative;
            overflow: hidden;
            backdrop-filter: blur(16px);
        }

        .metric-card::before {
            content: "";
            position: absolute;
            top: -35px;
            right: -30px;
            width: 110px;
            height: 110px;
            background: radial-gradient(circle, rgba(56,189,248,0.16), transparent 68%);
            pointer-events: none;
        }

        .metric-card::after {
            content: "";
            position: absolute;
            top: 0;
            left: 0;
            width: 100%;
            height: 4px;
            background: linear-gradient(90deg, #1e40af 0%, #2563eb 55%, #38bdf8 100%);
        }

        .metric-label {
            color: var(--text-3);
            font-size: 0.78rem;
            font-weight: 800;
            margin-bottom: 0.34rem;
            text-transform: uppercase;
            letter-spacing: 0.25px;
        }

        .metric-value {
            color: var(--text-1);
            font-size: 1.68rem;
            font-weight: 900;
            line-height: 1.04;
            margin-bottom: 0.16rem;
        }

        .metric-note {
            color: var(--text-3);
            font-size: 0.78rem;
            font-weight: 500;
        }

        .panel-card {
            background: rgba(255,255,255,0.78);
            border-radius: 22px;
            padding: 0.95rem;
            box-shadow: var(--shadow-sm);
            border: 1px solid rgba(255,255,255,0.46);
            margin-bottom: 0.85rem;
            backdrop-filter: blur(16px);
        }

        .section-title {
            font-size: 1rem;
            font-weight: 900;
            color: var(--text-1);
            margin-bottom: 0.08rem;
        }

        .section-subtitle {
            color: var(--text-3);
            font-size: 0.84rem;
            margin-bottom: 0.75rem;
            line-height: 1.5;
        }

        .small-note {
            color: var(--text-3);
            font-size: 0.79rem;
            padding-top: 0.2rem;
        }

        @media (max-width: 900px) {
            .hero-title { font-size: 0.88rem; }
            .hero-subtitle { font-size: 0.74rem; }
            .metric-value { font-size: 1.38rem; }
        
        .panel-card {
    background: transparent !important;
    border: none !important;
    box-shadow: none !important;
    padding: 0 !important;
    margin: 0 !important;
}
        }
    </style>
    """,
    unsafe_allow_html=True,
)


@st.cache_data(show_spinner=False)
def load_manning_master() -> Tuple[Optional[pd.DataFrame], Optional[str]]:
    try:
        spin_path = Path("Spinning.xlsx")
        wtt_path = Path("WTT.xlsx")

        if not spin_path.exists() or not wtt_path.exists():
            return None, "Spinning.xlsx or WTT.xlsx missing"

        spin_df = pd.read_excel(spin_path, sheet_name="Spinning")
        wtt_df = pd.read_excel(wtt_path, sheet_name="WTT")

        # --- FIX: align schema ---
        wtt_df = wtt_df.rename(columns={
            "HO_Scientific_Manpower": "BE_Scientific_Manpower",
            "HO_Final_Manpower": "BE_Final_Manpower",
        })

        # normalize Dept column
        for df in [spin_df, wtt_df]:
            if "Dept_Machine_Name" not in df.columns and "Department" in df.columns:
                df["Dept_Machine_Name"] = df["Department"]

        # --- STRICT schema check ---
        if set(spin_df.columns) != set(wtt_df.columns):
            return None, "Schema mismatch even after alignment"

        # merge
        source_df = pd.concat([spin_df, wtt_df], ignore_index=True)

        # enforce required columns
        for col in DISPLAY_COLUMNS:
            if col not in source_df.columns:
                source_df[col] = 0 if col in NUMERIC_COLUMNS else ""

        source_df["Excel_Row_No"] = range(2, len(source_df) + 2)

        return source_df, "Spinning + WTT merged"

    except Exception as e:
        return None, str(e)

def get_logo_path() -> Optional[str]:
    for candidate in FALLBACK_LOGO_PATHS:
        if candidate.exists():
            return str(candidate)
    return None


def safe_float(value, default=0.0):
    try:
        if value is None or value == "":
            return default
        if pd.isna(value):
            return default
        return float(value)
    except Exception:
        return default


def round_2(value):
    return round(safe_float(value), 2)


def excel_round(value, digits=0):
    return round(value, digits)


def excel_roundup(value, digits=0):
    factor = 10 ** digits
    return math.ceil(value * factor) / factor


def get_initial_tfo_data():
    data = [
        {"Count": "2/24 KW", "Customer": "Vapi", "Count2": 12.00, "Speed": 9000.00, "TPI": 8.50, "Utilization": 0.98, "Efficiency": 0.88, "Production Required / day Kgs": 9386.50, "TFO divisor": 240.00, "mpm": 850.00, "Eff": 0.85, "Machine divisor": 72.00},
        {"Count": "2/24 KW", "Customer": "Vapi", "Count2": 12.00, "Speed": 6000.00, "TPI": 8.50, "Utilization": 0.97, "Efficiency": 0.85, "Production Required / day Kgs": 0.00, "TFO divisor": 240.00, "mpm": 0.00, "Eff": 0.00, "Machine divisor": 72.00},
        {"Count": "2/30 CW bci Anjar", "Customer": "Anjar", "Count2": 15.00, "Speed": 10000.00, "TPI": 11.00, "Utilization": 0.98, "Efficiency": 0.90, "Production Required / day Kgs": 1408.00, "TFO divisor": 240.00, "mpm": 950.00, "Eff": 0.85, "Machine divisor": 72.00},
        {"Count": "2/20 KW", "Customer": "Vapi", "Count2": 10.00, "Speed": 8500.00, "TPI": 8.50, "Utilization": 0.98, "Efficiency": 0.87, "Production Required / day Kgs": 11968.00, "TFO divisor": 240.00, "mpm": 850.00, "Eff": 0.85, "Machine divisor": 72.00},
        {"Count": "2/20 KW BCI Anjar", "Customer": "Anjar", "Count2": 10.00, "Speed": 7000.00, "TPI": 8.50, "Utilization": 0.98, "Efficiency": 0.87, "Production Required / day Kgs": 0.00, "TFO divisor": 240.00, "mpm": 850.00, "Eff": 0.85, "Machine divisor": 72.00},
        {"Count": "2/12 K bci Anjar", "Customer": "Anjar", "Count2": 6.00, "Speed": 6000.00, "TPI": 6.50, "Utilization": 0.98, "Efficiency": 0.80, "Production Required / day Kgs": 1315.14, "TFO divisor": 240.00, "mpm": 800.00, "Eff": 0.85, "Machine divisor": 72.00},
        {"Count": "2/9 KC", "Customer": "Vapi", "Count2": 4.50, "Speed": 4000.00, "TPI": 6.50, "Utilization": 0.98, "Efficiency": 0.75, "Production Required / day Kgs": 3301.00, "TFO divisor": 240.00, "mpm": 700.00, "Eff": 0.85, "Machine divisor": 72.00},
        {"Count": "2/6 kc", "Customer": "Rugs", "Count2": 3.00, "Speed": 4000.00, "TPI": 6.50, "Utilization": 0.98, "Efficiency": 0.80, "Production Required / day Kgs": 0.00, "TFO divisor": 240.00, "mpm": 850.00, "Eff": 0.85, "Machine divisor": 72.00},
        {"Count": "2/6 kc", "Customer": "Rugs", "Count2": 3.00, "Speed": 3500.00, "TPI": 6.50, "Utilization": 0.98, "Efficiency": 0.80, "Production Required / day Kgs": 0.00, "TFO divisor": 240.00, "mpm": 850.00, "Eff": 0.85, "Machine divisor": 72.00},
        {"Count": "2/19 C hygro", "Customer": "Vapi", "Count2": 9.50, "Speed": 6000.00, "TPI": 6.50, "Utilization": 0.98, "Efficiency": 0.85, "Production Required / day Kgs": 0.00, "TFO divisor": 240.00, "mpm": 850.00, "Eff": 0.85, "Machine divisor": 72.00},
        {"Count": "9 com +80 PVA", "Customer": "Vapi", "Count2": 8.09, "Speed": 7000.00, "TPI": 13.00, "Utilization": 0.98, "Efficiency": 0.85, "Production Required / day Kgs": 0.00, "TFO divisor": 240.00, "mpm": 700.00, "Eff": 0.85, "Machine divisor": 72.00},
        {"Count": "DYED/ Others OS", "Customer": "Vapi", "Count2": 13.33, "Speed": 8000.00, "TPI": 15.00, "Utilization": 0.98, "Efficiency": 0.85, "Production Required / day Kgs": 2323.00, "TFO divisor": 240.00, "mpm": 850.00, "Eff": 0.85, "Machine divisor": 72.00},
        {"Count": "16+106 80 PVA OS", "Customer": "Vapi", "Count2": 13.33, "Speed": 8000.00, "TPI": 19.00, "Utilization": 0.98, "Efficiency": 0.90, "Production Required / day Kgs": 4287.00, "TFO divisor": 240.00, "mpm": 850.00, "Eff": 0.85, "Machine divisor": 72.00},
        {"Count": "13 hygro+80 PVA", "Customer": "Vapi", "Count2": 11.18, "Speed": 8500.00, "TPI": 13.00, "Utilization": 0.98, "Efficiency": 0.90, "Production Required / day Kgs": 3051.00, "TFO divisor": 240.00, "mpm": 850.00, "Eff": 0.85, "Machine divisor": 72.00},
        {"Count": "13 CC +80 PVA", "Customer": "Vapi", "Count2": 11.18, "Speed": 8500.00, "TPI": 13.00, "Utilization": 0.98, "Efficiency": 0.90, "Production Required / day Kgs": 4358.00, "TFO divisor": 240.00, "mpm": 850.00, "Eff": 0.85, "Machine divisor": 72.00},
        {"Count": "14 CC EGY +80 PVA", "Customer": "Vapi", "Count2": 11.18, "Speed": 7000.00, "TPI": 13.00, "Utilization": 0.98, "Efficiency": 0.90, "Production Required / day Kgs": 0.00, "TFO divisor": 240.00, "mpm": 850.00, "Eff": 0.85, "Machine divisor": 72.00},
        {"Count": "13 CC +80 PVA OS", "Customer": "Vapi", "Count2": 11.18, "Speed": 7000.00, "TPI": 13.00, "Utilization": 0.98, "Efficiency": 0.90, "Production Required / day Kgs": 0.00, "TFO divisor": 240.00, "mpm": 850.00, "Eff": 0.85, "Machine divisor": 72.00},
        {"Count": "4/2/6 K", "Customer": "Rugs", "Count2": 0.75, "Speed": 2000.00, "TPI": 2.80, "Utilization": 0.98, "Efficiency": 0.70, "Production Required / day Kgs": 2713.78, "TFO divisor": 80.00, "mpm": 400.00, "Eff": 0.70, "Machine divisor": 72.00},
    ]
    return pd.DataFrame(data)


if "tfo_input_df" not in st.session_state:
    st.session_state.tfo_input_df = get_initial_tfo_data()

if "full_spinning_editor_version" not in st.session_state:
    st.session_state.full_spinning_editor_version = 0


def calculate_upper_tfo_metrics(df: pd.DataFrame) -> pd.DataFrame:
    output_rows = []

    for _, row in df.iterrows():
        count2 = safe_float(row.get("Count2"))
        speed = safe_float(row.get("Speed"))
        tpi = safe_float(row.get("TPI"))
        efficiency = safe_float(row.get("Efficiency"))
        production_required_day = safe_float(row.get("Production Required / day Kgs"))
        tfo_divisor = safe_float(row.get("TFO divisor"), 240)
        mpm = safe_float(row.get("mpm"))
        eff_value = safe_float(row.get("Eff"))
        machine_divisor = safe_float(row.get("Machine divisor"), 72)

        production_required_month = production_required_day * 30 if production_required_day > 0 else 0.0

        if count2 > 0 and tpi > 0:
            production_per_drum_day = (
                (speed * 60 * 8 * efficiency * 2)
                / (tpi * 36 * 840 * count2 * 2.202)
            ) * 3
        else:
            production_per_drum_day = 0.0

        no_of_drums_required = (
            production_required_day / production_per_drum_day
            if production_per_drum_day > 0 else 0.0
        )

        no_of_tfo_required_shift = (
            no_of_drums_required / tfo_divisor
            if tfo_divisor > 0 else 0.0
        )

        if count2 > 0:
            kgs_per_drum_day = (
                (mpm * eff_value * 8 * 60 * 1.09)
                / (count2 * 840 * 2.202)
            ) * 3
        else:
            kgs_per_drum_day = 0.0

        no_of_drums = (
            production_required_day / kgs_per_drum_day
            if kgs_per_drum_day > 0 else 0.0
        )

        no_of_machines = (
            no_of_drums / machine_divisor
            if machine_divisor > 0 else 0.0
        )

        new_row = row.to_dict()
        new_row["Production Required / Month Kgs"] = round_2(production_required_month)
        new_row["Production per Drum/day"] = round_2(production_per_drum_day)
        new_row["No. of Drums Required"] = round_2(no_of_drums_required)
        new_row["No. of TFO Required / shift"] = round_2(no_of_tfo_required_shift)
        new_row["kgs/drum/day"] = round_2(kgs_per_drum_day)
        new_row["No. of Drums"] = round_2(no_of_drums)
        new_row["no. of machines"] = round_2(no_of_machines)

        output_rows.append(new_row)

    result_df = pd.DataFrame(output_rows)

    numeric_cols = [
        "Count2", "Speed", "TPI", "Utilization", "Efficiency",
        "Production Required / day Kgs", "Production Required / Month Kgs",
        "TFO divisor", "mpm", "Eff", "Machine divisor",
        "Production per Drum/day", "No. of Drums Required",
        "No. of TFO Required / shift", "kgs/drum/day",
        "No. of Drums", "no. of machines",
    ]

    for col in numeric_cols:
        if col in result_df.columns:
            result_df[col] = pd.to_numeric(result_df[col], errors="coerce").fillna(0).round(2)

    result_df["Upper_Row_No"] = range(2, len(result_df) + 2)
    return result_df


def build_upper_total_row(df: pd.DataFrame) -> pd.DataFrame:
    total_row = {
        "Count": "TOTAL",
        "Customer": "",
        "Count2": round_2(df["Count2"].sum()),
        "Speed": 0.00,
        "TPI": 0.00,
        "Utilization": 0.00,
        "Efficiency": 0.00,
        "Production Required / day Kgs": round_2(df["Production Required / day Kgs"].sum()),
        "Production Required / Month Kgs": round_2(df["Production Required / Month Kgs"].sum()),
        "Production per Drum/day": round_2(df["Production per Drum/day"].sum()),
        "No. of Drums Required": round_2(df["No. of Drums Required"].sum()),
        "No. of TFO Required / shift": round_2(df["No. of TFO Required / shift"].sum()),
        "mpm": 0.00,
        "Eff": 0.00,
        "kgs/drum/day": round_2(df["kgs/drum/day"].sum()),
        "No. of Drums": round_2(df["No. of Drums"].sum()),
        "no. of machines": round_2(df["no. of machines"].sum()),
        "TFO divisor": 0.00,
        "Machine divisor": 0.00,
    }
    return pd.DataFrame([total_row])


def split_shift(total_manpower, mode="three_shift"):
    total_manpower = int(round(total_manpower))

    if mode == "general":
        return total_manpower, 0, 0, 0

    shift_a = math.ceil(total_manpower / 3)
    shift_b = math.ceil((total_manpower - shift_a) / 2)
    shift_c = total_manpower - shift_a - shift_b
    return 0, shift_a, shift_b, shift_c


def build_tfo_row(
    sr_no,
    dept_machine_name,
    designation,
    formulas,
    scientific_manpower,
    general_shift,
    shift_a,
    shift_b,
    shift_c,
    reliever=0,
    operator_type="",
    contractors=0,
    company_associate=0,
    remarks="",
):
    return {
        "Location": "Vapi",
        "Business": "Spinning",
        "Section": "TFO",
        "Sr_No": sr_no,
        "Dept_Machine_Name": dept_machine_name,
        "Designation": designation,
        "Machine_Count": "",
        "Workload": "",
        "Formulas": formulas,
        "BE_Scientific_Manpower": round_2(scientific_manpower),
        "Operator_Type": operator_type,
        "Contractors": contractors,
        "Company_Associate": company_associate,
        "BE_Final_Manpower": round_2(scientific_manpower),
        "General_Shift": general_shift,
        "Shift_A": shift_a,
        "Shift_B": shift_b,
        "Shift_C": shift_c,
        "Reliever": reliever,
        "Remarks": remarks,
    }


def calculate_lower_tfo_manpower(upper_df: pd.DataFrame):
    sum_no_of_drums_total = round_2(upper_df["No. of Drums"].sum())
    sum_tfo_required_shift_total = round_2(upper_df["No. of TFO Required / shift"].sum())

    count_426k_mask = upper_df["Count"].astype(str).str.strip().str.upper() == "4/2/6 K"
    no_of_drums_426k = round_2(upper_df.loc[count_426k_mask, "No. of Drums"].sum())

    assembly_winding = excel_roundup((sum_no_of_drums_total / 36) * 3, 0)
    jumbo_assembly_winding = excel_round(excel_roundup(no_of_drums_426k, 0) / 16, 0) * 2
    tfo_operator = excel_round(sum_tfo_required_shift_total / 6, 0) * 3
    tfo_doffer = excel_roundup(sum_tfo_required_shift_total / 4, 0) * 3

    g1, a1, b1, c1 = split_shift(assembly_winding, "three_shift")
    g2, a2, b2, c2 = split_shift(jumbo_assembly_winding, "three_shift")
    g3, a3, b3, c3 = split_shift(tfo_operator, "three_shift")
    g4, a4, b4, c4 = split_shift(tfo_doffer, "three_shift")

    rows = [
        build_tfo_row(1, "Assembly winding", "operator", "ROUNDUP((SUM(T2:T18)/36)*3,0)", assembly_winding, g1, a1, b1, c1),
        build_tfo_row(2, "Jumbo Assembly Winding", "operator", "ROUND(ROUNDUP(T19,0)/16,0)*2", jumbo_assembly_winding, g2, a2, b2, c2),
        build_tfo_row(3, "TFO", "TFO Operator", "ROUND(SUM(N2:N18)/6,0)*3", tfo_operator, g3, a3, b3, c3),
        build_tfo_row(4, "TFO", "TFO Operator (Doffer)", "ROUNDUP(SUM(N2:N18)/4,0)*3", tfo_doffer, g4, a4, b4, c4),
        build_tfo_row(5, "Jumbo TFO", "TFO Operator", "2*3", 6, 0, 2, 2, 2),
        build_tfo_row(6, "Vijaylakshmi TFO", "TFO Operator", "2*3", 6, 0, 2, 2, 2),
        build_tfo_row(7, "Jobber", "Jobber", "1*3", 3, 0, 1, 1, 1),
        build_tfo_row(8, "Cone Carrier", "Cone carrier", "1*3", 3, 0, 1, 1, 1),
        build_tfo_row(9, "Cone Checker", "Cone checker", "2*3", 6, 0, 2, 2, 2),
        build_tfo_row(10, "Cone tipping", "", "1*3", 3, 0, 1, 1, 1),
        build_tfo_row(11, "Fork lift opt", "", "1*1", 1, 1, 0, 0, 0),
        build_tfo_row(12, "Packer", "", "4*3", 12, 0, 4, 4, 4),
        build_tfo_row(13, "Packing Jobber", "Jobber", "0*3", 0, 0, 0, 0, 0),
        build_tfo_row(15, "DEO", "DEO", "1*1", 1, 1, 0, 0, 0),
        build_tfo_row(16, "House Keeper", "Contractors", "2*3", 6, 0, 2, 2, 2),
    ]

    lower_df = pd.DataFrame(rows)

    for col in NUMERIC_COLUMNS:
        if col in lower_df.columns:
            lower_df[col] = pd.to_numeric(lower_df[col], errors="coerce").fillna(0).round(2)

    return lower_df, {
        "sum_no_of_drums_total": round_2(sum_no_of_drums_total),
        "sum_tfo_required_shift_total": round_2(sum_tfo_required_shift_total),
        "no_of_drums_426k": round_2(no_of_drums_426k),
    }


def sanitize_editor_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    safe_df = df.copy()
    safe_df = safe_df.replace([math.inf, -math.inf], pd.NA)

    for col in NUMERIC_COLUMNS:
        if col in safe_df.columns:
            safe_df[col] = pd.to_numeric(safe_df[col], errors="coerce")

    for col in TEXT_COLUMNS:
        if col in safe_df.columns:
            safe_df[col] = safe_df[col].fillna("").astype(str)

    return safe_df


def normalize_compare_df(df: pd.DataFrame) -> pd.DataFrame:
    compare_df = df.copy()

    for col in compare_df.columns:
        if col in NUMERIC_COLUMNS or col == "BE_Scientific_Manpower":
            compare_df[col] = pd.to_numeric(compare_df[col], errors="coerce").fillna(0).round(4)
        else:
            compare_df[col] = compare_df[col].fillna("").astype(str).str.strip()

    return compare_df.reset_index(drop=True)


def dataframes_equal_for_ui(left_df: pd.DataFrame, right_df: pd.DataFrame) -> bool:
    return normalize_compare_df(left_df).equals(normalize_compare_df(right_df))


def get_initial_full_spinning_df(source_df: pd.DataFrame, lower_tfo_df: pd.DataFrame) -> pd.DataFrame:
    non_tfo_df = source_df.loc[source_df["Section"].astype(str).str.upper() != "TFO"].copy()
    original_tfo_rows = source_df.loc[source_df["Section"].astype(str).str.upper() == "TFO", "Excel_Row_No"].tolist()

    dynamic_tfo_df = lower_tfo_df.copy()
    if len(original_tfo_rows) >= len(dynamic_tfo_df):
        dynamic_tfo_df["Excel_Row_No"] = original_tfo_rows[: len(dynamic_tfo_df)]
    else:
        start_row = int(source_df["Excel_Row_No"].max()) + 1
        dynamic_tfo_df["Excel_Row_No"] = list(range(start_row, start_row + len(dynamic_tfo_df)))

    non_tfo_df["Row_Key"] = [f"BASE_{row_no}" for row_no in non_tfo_df["Excel_Row_No"]]
    dynamic_tfo_df["Row_Key"] = [f"TFO_{i + 1}" for i in range(len(dynamic_tfo_df))]

    full_df = pd.concat([non_tfo_df, dynamic_tfo_df], ignore_index=True)
    return sanitize_editor_dataframe(full_df)


def assign_tfo_row_metadata(source_df: pd.DataFrame, lower_tfo_df: pd.DataFrame) -> pd.DataFrame:
    prepared_tfo_df = lower_tfo_df.copy()

    original_tfo_rows = source_df.loc[
        source_df["Section"].astype(str).str.upper() == "TFO",
        "Excel_Row_No",
    ].tolist()

    if len(original_tfo_rows) >= len(prepared_tfo_df):
        prepared_tfo_df["Excel_Row_No"] = original_tfo_rows[: len(prepared_tfo_df)]
    else:
        start_row = int(source_df["Excel_Row_No"].max()) + 1
        prepared_tfo_df["Excel_Row_No"] = list(range(start_row, start_row + len(prepared_tfo_df)))

    prepared_tfo_df["Row_Key"] = [f"TFO_{i + 1}" for i in range(len(prepared_tfo_df))]
    return prepared_tfo_df


def sum_spinning_reference(current_df: pd.DataFrame, col_letter: str, start_row: int, end_row: int) -> float:
    column_map = {"K": "BE_Scientific_Manpower"}
    column_name = column_map.get(col_letter.upper())

    if not column_name or column_name not in current_df.columns:
        return 0.0

    mask = current_df["Excel_Row_No"].between(start_row, end_row, inclusive="both")
    values = pd.to_numeric(current_df.loc[mask, column_name], errors="coerce").fillna(0)
    return float(values.sum())


def upper_tfo_reference_value(upper_df: pd.DataFrame, col_letter: str, row_no: int) -> float:
    column_map = {
        "N": "No. of TFO Required / shift",
        "T": "No. of Drums",
    }
    column_name = column_map.get(col_letter.upper())

    if not column_name or column_name not in upper_df.columns:
        return 0.0

    matched = upper_df.loc[upper_df["Upper_Row_No"] == row_no, column_name]
    if matched.empty:
        return 0.0

    return safe_float(matched.iloc[0], 0.0)


def evaluate_formula(formula_text: str, current_df: pd.DataFrame, upper_tfo_df: pd.DataFrame):
    if formula_text is None:
        return None

    expression = str(formula_text).strip()
    if expression == "" or expression.lower() == "nan":
        return None

    expression = expression.lstrip("=")
    expression = expression.replace("^", "**")
    expression = re.sub(r"(\d+(?:\.\d+)?)\s*%", lambda match: f"({match.group(1)}/100)", expression)

    range_pattern = re.compile(r"SUM\(\+?([A-Z]+)(\d+):([A-Z]+)(\d+)\)", re.IGNORECASE)

    def range_replacer(match):
        start_col = match.group(1).upper()
        end_col = match.group(3).upper()
        start_row = int(match.group(2))
        end_row = int(match.group(4))

        if start_col != end_col:
            return "0"

        if start_col == "K":
            return str(sum_spinning_reference(current_df, start_col, start_row, end_row))

        if start_col in {"N", "T"}:
            return str(
                sum(
                    upper_tfo_reference_value(upper_tfo_df, start_col, row_no)
                    for row_no in range(start_row, end_row + 1)
                )
            )

        return "0"

    expression = range_pattern.sub(range_replacer, expression)

    single_ref_pattern = re.compile(r"(?<![A-Z])([KNT])(\d+)(?![A-Z])", re.IGNORECASE)

    def single_ref_replacer(match):
        col_letter = match.group(1).upper()
        row_no = int(match.group(2))

        if col_letter == "K":
            return str(sum_spinning_reference(current_df, col_letter, row_no, row_no))

        return str(upper_tfo_reference_value(upper_tfo_df, col_letter, row_no))

    expression = single_ref_pattern.sub(single_ref_replacer, expression)
    expression = re.sub(r"\bROUNDUP\s*\(", "excel_roundup(", expression, flags=re.IGNORECASE)
    expression = re.sub(r"\bROUND\s*\(", "excel_round(", expression, flags=re.IGNORECASE)

    safe_namespace = {
        "excel_round": excel_round,
        "excel_roundup": excel_roundup,
        "math": math,
        "abs": abs,
        "min": min,
        "max": max,
    }

    try:
        result = eval(expression, {"__builtins__": {}}, safe_namespace)
        result = safe_float(result, 0.0)

        if math.isfinite(result):
            return round(result, 2)

        return 0.0
    except Exception:
        return None


def recalculate_scientific_manpower(full_df: pd.DataFrame, upper_tfo_df: pd.DataFrame) -> pd.DataFrame:
    recalculated_df = full_df.copy()
    recalculated_df = recalculated_df.sort_values("Excel_Row_No").reset_index(drop=True)

    for index, row in recalculated_df.iterrows():
        formula_value = row.get("Formulas", "")
        evaluated_value = evaluate_formula(formula_value, recalculated_df, upper_tfo_df)

        if evaluated_value is not None:
            recalculated_df.at[index, "BE_Scientific_Manpower"] = round_2(evaluated_value)
        else:
            existing_value = row.get("BE_Scientific_Manpower")

            if pd.isna(existing_value) or existing_value == "":
                recalculated_df.at[index, "BE_Scientific_Manpower"] = pd.NA
            else:
                recalculated_df.at[index, "BE_Scientific_Manpower"] = round_2(existing_value)

    return sanitize_editor_dataframe(recalculated_df)


def build_summary_table(full_spinning_df: pd.DataFrame) -> pd.DataFrame:
    summary_df = full_spinning_df.groupby(
        ["Location", "Business", "Section"],
        as_index=False,
        dropna=False
    )["BE_Final_Manpower"].sum()

    summary_df["BE_Final_Manpower"] = (
        pd.to_numeric(summary_df["BE_Final_Manpower"], errors="coerce")
        .fillna(0)
        .round(0)
        .astype(int)
    )

    return summary_df.sort_values(["Location", "Business", "Section"]).reset_index(drop=True)


def apply_editor_changes(master_df: pd.DataFrame, edited_view_df: pd.DataFrame) -> pd.DataFrame:
    updated_df = master_df.copy()

    editable_plus_key = ["Row_Key"] + EDITABLE_COLUMNS
    changed_df = edited_view_df[editable_plus_key].copy()
    changed_df = sanitize_editor_dataframe(changed_df)

    updated_df = (
        updated_df
        .drop(columns=EDITABLE_COLUMNS, errors="ignore")
        .merge(changed_df, on="Row_Key", how="left")
    )

    return sanitize_editor_dataframe(updated_df)


def rebuild_full_spinning_with_tfo(
    source_df: pd.DataFrame,
    base_full_df: pd.DataFrame,
    upper_tfo_df: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame, dict]:
    fresh_lower_tfo_df, driver_values = calculate_lower_tfo_manpower(upper_tfo_df)
    fresh_lower_tfo_df["BE_Final_Manpower"] = fresh_lower_tfo_df["BE_Scientific_Manpower"]
    fresh_lower_tfo_df = assign_tfo_row_metadata(source_df, fresh_lower_tfo_df)

    non_tfo_df = base_full_df.loc[
        base_full_df["Section"].astype(str).str.upper() != "TFO"
    ].copy()

    updated_full_df = pd.concat([non_tfo_df, fresh_lower_tfo_df], ignore_index=True)
    updated_full_df = sanitize_editor_dataframe(updated_full_df)
    updated_full_df = recalculate_scientific_manpower(updated_full_df, upper_tfo_df)

    tfo_mask = updated_full_df["Section"].astype(str).str.upper() == "TFO"
    updated_full_df.loc[tfo_mask, "BE_Final_Manpower"] = updated_full_df.loc[
        tfo_mask, "BE_Scientific_Manpower"
    ]

    updated_full_df = sanitize_editor_dataframe(updated_full_df)
    current_lower_df = updated_full_df.loc[tfo_mask, DISPLAY_COLUMNS].copy()

    return updated_full_df, current_lower_df, driver_values


def create_download_workbook(
    summary_df: pd.DataFrame,
    full_spinning_df: pd.DataFrame,
    upper_tfo_df: pd.DataFrame,
    lower_tfo_df: pd.DataFrame,
) -> bytes:
    output = BytesIO()

    export_spinning_df = full_spinning_df[DISPLAY_COLUMNS].copy()
    export_tfo_upper_df = upper_tfo_df.drop(columns=["Upper_Row_No"], errors="ignore").copy()
    export_tfo_lower_df = lower_tfo_df[DISPLAY_COLUMNS].copy()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        export_spinning_df.to_excel(writer, sheet_name="Spinning", index=False)

        tfo_sheet_name = "TFO"
        export_tfo_upper_df.to_excel(writer, sheet_name=tfo_sheet_name, index=False, startrow=1)
        lower_start_row = len(export_tfo_upper_df) + 5
        export_tfo_lower_df.to_excel(writer, sheet_name=tfo_sheet_name, index=False, startrow=lower_start_row)

        tfo_sheet = writer.sheets[tfo_sheet_name]
        summary_sheet = writer.sheets["Summary"]
        spinning_sheet = writer.sheets["Spinning"]

        header_fill = PatternFill("solid", fgColor="0F4C81")
        header_font = Font(color="FFFFFF", bold=True)
        title_fill = PatternFill("solid", fgColor="D9EAF7")
        title_font = Font(color="0F172A", bold=True)

        tfo_sheet["A1"] = "Upper TFO Production Table"
        tfo_sheet["A1"].font = title_font
        tfo_sheet["A1"].fill = title_fill

        lower_title_row = lower_start_row + 1
        tfo_sheet.cell(row=lower_title_row, column=1, value="Lower TFO Manpower Table")
        tfo_sheet.cell(row=lower_title_row, column=1).font = title_font
        tfo_sheet.cell(row=lower_title_row, column=1).fill = title_fill

        for sheet in [summary_sheet, spinning_sheet, tfo_sheet]:
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    if cell.row in {1, 2, lower_start_row + 1, lower_start_row + 2}:
                        if isinstance(cell.value, str) and cell.value not in {
                            "Upper TFO Production Table",
                            "Lower TFO Manpower Table",
                        }:
                            cell.fill = header_fill
                            cell.font = header_font

            for column_cells in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    cell_value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, len(cell_value))
                sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 34)

    output.seek(0)
    return output.getvalue()


def render_metric_card(title: str, value: str, note: str):
    st.markdown(
        f"""
        <div class="metric-card">
            <div class="metric-label">{title}</div>
            <div class="metric-value">{value}</div>
            <div class="metric-note">{note}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


try:
    import os
    import pandas as pd
    from schema_utils import align_and_validate_schemas

    DATA_DIR = "data"
    all_dfs = []

    for location in os.listdir(DATA_DIR):

        location_path = os.path.join(DATA_DIR, location)

        if not os.path.isdir(location_path):
            continue

        spinning_path = os.path.join(location_path, "Spinning.xlsx")
        rugs_path = os.path.join(location_path, "Rugs.xlsx")
        wtt_path = os.path.join(location_path, "WTT.xlsx")

        # -------------------------
        # CASE 1: Vapi-like (3 files)
        # -------------------------
        if os.path.exists(spinning_path) and os.path.exists(rugs_path):

            spin_df, wtt_df, rugs_df = align_and_validate_schemas(
                spinning_path,
                wtt_path,
                rugs_path
            )

            dfs = [spin_df, wtt_df, rugs_df]

        # -------------------------
        # CASE 2: Single file (Anjar etc.)
        # -------------------------
        else:
            # pick any excel file in folder
            excel_files = [f for f in os.listdir(location_path) if f.endswith(".xlsx")]

            dfs = []
            for file in excel_files:
                file_path = os.path.join(location_path, file)

                try:
                    xls = pd.ExcelFile(file_path)
                    sheet_name = os.path.splitext(file)[0]  # sheet = filename

                    df = pd.read_excel(xls, sheet_name=sheet_name)

                    from schema_utils import normalize_columns, compute_n_shifts
                    df = normalize_columns(df)
                    df = compute_n_shifts(df)

                    # df["Business"] = sheet_name
                    dfs.append(df)

                except Exception as e:
                    print(f"Skipping {file_path}: {e}")

        # -------------------------
        # Add location + collect
        # -------------------------
        for df in dfs:
            if df is not None and not df.empty:
                df["Location"] = location
                all_dfs.append(df)

    # -------------------------
    # FINAL DF
    # -------------------------
    source_spinning_df = pd.concat(all_dfs, ignore_index=True)

    # st.write("Locations present:", source_spinning_df["Location"].unique())

    source_spinning_df["Location"] = source_spinning_df["Location"].astype(str).str.strip()

    source_spinning_df["Excel_Row_No"] = range(2, len(source_spinning_df) + 2)

except Exception as e:
    st.error(str(e))
    st.stop()

current_upper_df = calculate_upper_tfo_metrics(st.session_state.tfo_input_df)

if "full_spinning_df" not in st.session_state:
    initial_lower_tfo_df, _ = calculate_lower_tfo_manpower(current_upper_df)
    st.session_state.full_spinning_df = get_initial_full_spinning_df(source_spinning_df, initial_lower_tfo_df)
    st.session_state.full_spinning_df = recalculate_scientific_manpower(
        st.session_state.full_spinning_df,
        current_upper_df,
    )
    tfo_mask_init = st.session_state.full_spinning_df["Section"].astype(str).str.upper() == "TFO"
    st.session_state.full_spinning_df.loc[tfo_mask_init, "BE_Final_Manpower"] = (
        st.session_state.full_spinning_df.loc[tfo_mask_init, "BE_Scientific_Manpower"]
    )

st.session_state.full_spinning_df = recalculate_scientific_manpower(
    st.session_state.full_spinning_df,
    current_upper_df
)

full_spinning_df = st.session_state.full_spinning_df.copy()
summary_df = build_summary_table(full_spinning_df)
logo_path = get_logo_path()

st.markdown('<div class="hero-shell">', unsafe_allow_html=True)
hero_col_1, hero_col_2 = st.columns([0.5, 3.5], vertical_alignment="center")

with hero_col_1:
    if logo_path:
        # st.markdown('<div class="hero-logo-card">', unsafe_allow_html=True)
        st.image(logo_path, width=200)
        st.markdown("</div>", unsafe_allow_html=True)

with hero_col_2:
    # st.markdown('<div class="hero-badge">Welspun Smart Operations</div>', unsafe_allow_html=True)
    st.markdown('<div class="hero-gradient-title">Manpower Recommendation Engine</div>', unsafe_allow_html=True)
    st.markdown(
        """
        <div class="hero-subtitle">
            Enabling optimized manpower allocation, using a production-driven recommendation engine
        </div>
        """,
        unsafe_allow_html=True,
    )

st.markdown("</div>", unsafe_allow_html=True)

st.markdown('<div class="premium-strip">', unsafe_allow_html=True)
kpi_col_1, kpi_col_2, kpi_col_3, kpi_col_4 = st.columns(4)

with kpi_col_1:
    render_metric_card(
        "Total Final Manpower",
        f"{int(round(pd.to_numeric(full_spinning_df['BE_Final_Manpower'], errors='coerce').fillna(0).sum())):,}",
        "Across all sections",
    )

with kpi_col_2:
    render_metric_card(
        "Scientific Manpower",
        f"{pd.to_numeric(full_spinning_df['BE_Scientific_Manpower'], errors='coerce').fillna(0).sum():,.2f}",
        "Driven by formulas",
    )

with kpi_col_3:
    render_metric_card(
        "Sections Covered",
        f"{full_spinning_df['Section'].nunique():,}",
        "Operational coverage",
    )

with kpi_col_4:
    render_metric_card(
        "Locations",
        f"{full_spinning_df['Location'].nunique():,}",
        "Active scope",
    )

st.markdown("</div>", unsafe_allow_html=True)

def render_location_tab(location_name, full_df, source_df, wtt_path=None):
    # if location_name.upper() in ["WFL","HYDERABAD"]:
    #     return

    # display_name = "Hyderabad" if location_name.upper() == "WFL" else location_name
    # full_df = full_df.copy()
    # # current change
    # if location_name.upper() == "ANJAR":
    #     sub_tabs = st.tabs(["Manpower"])
    # elif location_name.upper() == "Hyderabad":
    #     sub_tabs = st.tabs(["Empty"])
    # else:
    #     if location_name.upper() == "ANJAR":
    #         # sub_tabs = st.tabs(["Manpower"])
    #          st.link_button("Go to Anjar", "https://spinning.streamlit.app/")
    #     elif location_name.upper() == "Hyderabad":
    #         sub_tabs = st.tabs(["Empty"])
    #     else:
    #         if location_name.upper() in ["ANJAR", "WFL"]:
    #             sub_tabs = st.tabs(["Manpower"])
    #         else:
    #             sub_tabs = st.tabs(["Manpower", "Rugs", "Spinning", "TT"])

    

# -----------------
# TAB CREATION
# -----------------

    loc = location_name.upper()
    if loc == "ANJAR":
        sub_tabs = st.tabs(["Anjar"])
    elif loc in ["WFL", "HYDERABAD"]:
        return
    else:
        sub_tabs = st.tabs(["Manpower", "Rugs", "Spinning", "WTT"])

    # -----------------
    # ANJAR (button only)
    # -----------------
    if loc == "ANJAR":
        with sub_tabs[0]:
            # st.markdown("### Anjar Module")
            st.link_button("Go to Anjar", "https://welspunanjar.streamlit.app/")
        return

    # -----------------
    # VAPI MANPOWER
    # -----------------
    with sub_tabs[0]:
    #     if location_name.upper() == "ANJAR":
            # df = full_df[full_df["Location"].str.upper() == "ANJAR"]
            # st.dataframe(df[DISPLAY_COLUMNS], width="stretch", hide_index=True)
            # return

        # if location_name.upper() == "WFL":
        #     return

        df = full_df[full_df["Location"].str.upper() == location_name.upper()]
        st.dataframe(df[DISPLAY_COLUMNS], width="stretch", hide_index=True)

    # -----------------
    # RUGS
    # -----------------
    if location_name.upper() in ["ANJAR", "WFL"]:
        return
    with sub_tabs[1]:
        # df = full_df[
        #     (full_df["Location"].str.upper() == location_name.upper()) &
        #     (full_df["Business"].str.upper() == "RUGS")
        # ]
        # st.dataframe(df[DISPLAY_COLUMNS], width="stretch", hide_index=True)
         if location_name.upper() == "VAPI":
            st.link_button("Go to Rugs", "https://example.com/rugs")
            

    # -----------------
    # SPINNING
    # -----------------
    if location_name.upper() in ["ANJAR", "WFL"]:
        return
    with sub_tabs[2]:

        if location_name.upper() == "VAPI":
            st.link_button("Go to Spinning", "https://spinning.streamlit.app/")
            

        # ✅ define subtabs
    #     spin_sub1, spin_sub2 = st.tabs(["Main", "TFO"])

    #     # Main
    #       # -----------------
    #     # MAIN SPINNING
    #     # -----------------
    #     with spin_sub1:
    #         df = full_df[
    #     (full_df["Location"].str.upper() == location_name.upper()) &
    #     (full_df["Business"].str.upper() == "SPINNING")
    # ]
    #         st.dataframe(df[DISPLAY_COLUMNS], width="stretch", hide_index=True)


    #     # TFO (ONLY for Vapi)

    #     with spin_sub2:

    #         if location_name.upper() == "VAPI":

    #             # ✅ DIRECTLY paste your TFO code here (no extra with)

    #             # st.markdown('<div class="panel-card">', unsafe_allow_html=True)
    #             st.markdown('<div class="section-title">TFO planning and manpower engine</div>', unsafe_allow_html=True)
    #             st.markdown(
    #                 '<div class="section-subtitle">Editable TFO production inputs with automatic roll-through into the entire spinning table and final summary.</div>',
    #                 unsafe_allow_html=True,
    #             )

    #             input_columns = [
    #                 "Count",
    #                 "Customer",
    #                 "Count2",
    #                 "Speed",
    #                 "TPI",
    #                 "Utilization",
    #                 "Efficiency",
    #                 "Production Required / day Kgs",
    #                 "TFO divisor",
    #                 "mpm",
    #                 "Eff",
    #                 "Machine divisor",
    #             ]

    #             edited_tfo_input_df = st.data_editor(
    #                 st.session_state.tfo_input_df[input_columns],
    #                 width="stretch",
    #                 num_rows="dynamic",
    #                 key="tfo_editor",
    #                 hide_index=True,
    #                 column_config={
    #                     "Count": st.column_config.TextColumn("Count"),
    #                     "Customer": st.column_config.TextColumn("Customer"),
    #                     "Count2": st.column_config.NumberColumn("Count2", format="%.2f"),
    #                     "Speed": st.column_config.NumberColumn("Speed", format="%.2f"),
    #                     "TPI": st.column_config.NumberColumn("TPI", format="%.2f"),
    #                     "Utilization": st.column_config.NumberColumn("Utilization", format="%.2f"),
    #                     "Efficiency": st.column_config.NumberColumn("Efficiency", format="%.2f"),
    #                     "Production Required / day Kgs": st.column_config.NumberColumn("Production Required / day Kgs", format="%.2f"),
    #                     "TFO divisor": st.column_config.NumberColumn("TFO divisor", format="%.2f"),
    #                     "mpm": st.column_config.NumberColumn("mpm", format="%.2f"),
    #                     "Eff": st.column_config.NumberColumn("Eff", format="%.2f"),
    #                     "Machine divisor": st.column_config.NumberColumn("Machine divisor", format="%.2f"),
    #                 },
    #             )

    #             for col in [
    #                 "Count2",
    #                 "Speed",
    #                 "TPI",
    #                 "Utilization",
    #                 "Efficiency",
    #                 "Production Required / day Kgs",
    #                 "TFO divisor",
    #                 "mpm",
    #                 "Eff",
    #                 "Machine divisor",
    #             ]:
    #                 edited_tfo_input_df[col] = pd.to_numeric(
    #                     edited_tfo_input_df[col],
    #                     errors="coerce",
    #                 ).fillna(0).round(2)

    #             st.session_state.tfo_input_df = edited_tfo_input_df.copy()

    #             current_upper_df = calculate_upper_tfo_metrics(edited_tfo_input_df)
    #             current_upper_total_df = build_upper_total_row(current_upper_df)
    #             current_upper_final_df = pd.concat(
    #                 [current_upper_df.drop(columns=["Upper_Row_No"], errors="ignore"), current_upper_total_df],
    #                 ignore_index=True,
    #             )

    #             updated_full_df, current_lower_df, current_driver_values = rebuild_full_spinning_with_tfo(
    #                 source_df=source_spinning_df,
    #                 base_full_df=st.session_state.full_spinning_df,
    #                 upper_tfo_df=current_upper_df,
    #             )

    #             st.session_state.full_spinning_df = updated_full_df
    #             full_df = updated_full_df.copy()

    #             tfo_metric_1, tfo_metric_2, tfo_metric_3, tfo_metric_4 = st.columns(4)

    #             with tfo_metric_1:
    #                 render_metric_card(
    #                     "Total No. of Drums",
    #                     f"{current_driver_values['sum_no_of_drums_total']:.2f}",
    #                     "Calculated from inputs",
    #                 )

    #             with tfo_metric_2:
    #                 render_metric_card(
    #                     "TFO Required / Shift",
    #                     f"{current_driver_values['sum_tfo_required_shift_total']:.2f}",
    #                     "Based on divisor logic",
    #                 )

    #             with tfo_metric_3:
    #                 render_metric_card(
    #                     "Drums for 4/2/6 K",
    #                     f"{current_driver_values['no_of_drums_426k']:.2f}",
    #                     "Used in Jumbo Assembly",
    #                 )

    #             with tfo_metric_4:
    #                 render_metric_card(
    #                     "Lower TFO Final Manpower",
    #                     f"{int(round(pd.to_numeric(current_lower_df['BE_Final_Manpower'], errors='coerce').fillna(0).sum())):,}",
    #                     "Current TFO rows",
    #                 )

    #             st.markdown("#### Upper TFO Production Table")

    #             upper_display_columns = [
    #                 "Count",
    #                 "Customer",
    #                 "Count2",
    #                 "Speed",
    #                 "TPI",
    #                 "Utilization",
    #                 "Efficiency",
    #                 "Production per Drum/day",
    #                 "Production Required / day Kgs",
    #                 "Production Required / Month Kgs",
    #                 "No. of Drums Required",
    #                 "No. of TFO Required / shift",
    #                 "mpm",
    #                 "Eff",
    #                 "kgs/drum/day",
    #                 "No. of Drums",
    #                 "no. of machines",
    #             ]

    #             st.dataframe(
    #                 current_upper_final_df[upper_display_columns],
    #                 width="stretch",
    #                 hide_index=True,
    #             )

    #             st.markdown("#### Lower TFO Manpower Table")
    #             st.dataframe(
    #                 current_lower_df,
    #                 width="stretch",
    #                 hide_index=True,
    #                 column_config={
    #                     "BE_Scientific_Manpower": st.column_config.NumberColumn("BE_Scientific_Manpower", format="%.2f"),
    #                     "BE_Final_Manpower": st.column_config.NumberColumn("BE_Final_Manpower", format="%.2f"),
    #                 },
    #             )

    #             with st.expander("Formula logic"):
    #                 st.markdown(
    #                     f"""
    #                     **Upper table**
    #                     - Production Required / Month Kgs = Production Required / day Kgs × 30
    #                     - Production per Drum/day = `((Speed × 60 × 8 × Efficiency × 2) / (TPI × 36 × 840 × Count2 × 2.202)) × 3`
    #                     - No. of Drums Required = Production Required / day Kgs / Production per Drum/day
    #                     - No. of TFO Required / shift = No. of Drums Required / TFO divisor
    #                     - kgs/drum/day = `((mpm × Eff × 8 × 60 × 1.09) / (Count2 × 840 × 2.202)) × 3`
    #                     - No. of Drums = Production Required / day Kgs / kgs/drum/day
    #                     - no. of machines = No. of Drums / Machine divisor

    #                     **Lower table**
    #                     - Assembly winding = `ROUNDUP((SUM(T2:T18)/36)*3,0)`
    #                     - Jumbo Assembly Winding = `ROUND(ROUNDUP(T19,0)/16,0)*2`
    #                     - TFO Operator = `ROUND(SUM(N2:N18)/6,0)*3`
    #                     - TFO Operator (Doffer) = `ROUNDUP(SUM(N2:N18)/4,0)*3`

    #                     **Current driver values**
    #                     - Total No. of Drums = {current_driver_values['sum_no_of_drums_total']:.2f}
    #                     - Total No. of TFO Required / shift = {current_driver_values['sum_tfo_required_shift_total']:.2f}
    #                     - No. of Drums for 4/2/6 K = {current_driver_values['no_of_drums_426k']:.2f}
    #                     """
    #                 )

    #             action_col_1, action_col_2 = st.columns(2)

    #             with action_col_1:
    #                 if st.button("Reset TFO Table"):
    #                     st.session_state.tfo_input_df = get_initial_tfo_data()
    #                     reset_upper_df = calculate_upper_tfo_metrics(st.session_state.tfo_input_df)
    #                     reset_full_df, _, _ = rebuild_full_spinning_with_tfo(
    #                         source_df=source_spinning_df,
    #                         base_full_df=st.session_state.full_spinning_df,
    #                         upper_tfo_df=reset_upper_df,
    #                     )
    #                     st.session_state.full_spinning_df = reset_full_df
    #                     st.session_state.full_spinning_editor_version += 1
    #                     st.rerun()

    #             with action_col_2:
    #                 if st.button("Reset Full Spinning Table from Source"):
    #                     fresh_upper_df = calculate_upper_tfo_metrics(st.session_state.tfo_input_df)
    #                     fresh_lower_df, _ = calculate_lower_tfo_manpower(fresh_upper_df)
    #                     st.session_state.full_spinning_df = get_initial_full_spinning_df(source_spinning_df, fresh_lower_df)
    #                     st.session_state.full_spinning_df = recalculate_scientific_manpower(
    #                         st.session_state.full_spinning_df,
    #                         fresh_upper_df,
    #                     )
    #                     tfo_mask = st.session_state.full_spinning_df["Section"].astype(str).str.upper() == "TFO"
    #                     st.session_state.full_spinning_df.loc[tfo_mask, "BE_Final_Manpower"] = (
    #                         st.session_state.full_spinning_df.loc[tfo_mask, "BE_Scientific_Manpower"]
    #                     )
    #                     st.session_state.full_spinning_editor_version += 1
    #                     st.rerun()

    #             st.markdown("</div>", unsafe_allow_html=True)


    #             # --- STEP 4: Recalculate scientific manpower ---
    #             # --- STEP 4: Recalculate scientific manpower ---
    #             current_upper_df = calculate_upper_tfo_metrics(st.session_state.tfo_input_df)

    #             st.session_state.full_spinning_df = recalculate_scientific_manpower(
    #                 st.session_state.full_spinning_df,
    #                 current_upper_df
    #             )

    #             pass
    #         else:
    #             st.info("TFO not configured for this location.")

    # -----------------
    # WTT
    # -----------------
    if location_name.upper() in ["ANJAR", "WFL"]:
        return
    with sub_tabs[3]:

        if location_name.upper() == "VAPI":
            st.link_button("Go to TT", "https://vttwelspun.streamlit.app/")
            

        # df = full_df[
        #     (full_df["Location"].str.upper() == location_name.upper()) &
        #     (full_df["Business"].str.upper() == "WTT")
        # ]
        # st.dataframe(df[DISPLAY_COLUMNS], width="stretch", hide_index=True)

summary_tab, vapi_tab, anjar_tab, wfl_tab = st.tabs(
    ["Summary", "Vapi", "Anjar", "Hyderabad"]
)

with summary_tab:
    # st.markdown('<div class="panel-card">', unsafe_allow_html=True)

    st.markdown('<div class="section-title">Executive filters</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="section-subtitle">Use the filters below to focus the summary view without changing the underlying data.</div>',
        unsafe_allow_html=True,
    )

    filter_col_1, filter_col_2, filter_col_3 = st.columns([1, 1, 2])

    with filter_col_1:
        selected_location = st.selectbox(
            "Location",
            options=sorted(summary_df["Location"].dropna().unique().tolist()),
        )

    with filter_col_2:
        business_options = sorted(
            summary_df.loc[summary_df["Location"] == selected_location, "Business"]
            .dropna()
            .unique()
            .tolist()
        )
        selected_business = st.selectbox("Business", options=business_options)

    with filter_col_3:
        section_options = sorted(
            summary_df.loc[
                (summary_df["Location"] == selected_location)
                & (summary_df["Business"] == selected_business),
                "Section",
            ]
            .dropna()
            .unique()
            .tolist()
        )
        selected_sections = st.multiselect(
            "Section filter",
            options=section_options,
            default=section_options,
        )

    st.markdown("</div>", unsafe_allow_html=True)

    filtered_summary_df = summary_df.loc[
        (summary_df["Location"] == selected_location)
        & (summary_df["Business"] == selected_business)
    ].copy()

    if selected_sections:
        filtered_summary_df = filtered_summary_df.loc[
            filtered_summary_df["Section"].isin(selected_sections)
        ].copy()
    else:
        filtered_summary_df = filtered_summary_df.iloc[0:0].copy()

    summary_total = int(
        round(pd.to_numeric(filtered_summary_df["BE_Final_Manpower"], errors="coerce").fillna(0).sum())
    )
    summary_sections = int(filtered_summary_df["Section"].nunique())

    metric_col_1, metric_col_2, metric_col_3 = st.columns(3)

    with metric_col_1:
        render_metric_card("Filtered Final Manpower", f"{summary_total:,}", "Visible selection")

    with metric_col_2:
        render_metric_card("Visible Sections", f"{summary_sections:,}", "Current output")

    with metric_col_3:
        top_section = "-"
        if not filtered_summary_df.empty:
            top_row = filtered_summary_df.sort_values("BE_Final_Manpower", ascending=False).iloc[0]
            top_section = f"{top_row['Section']} ({int(top_row['BE_Final_Manpower'])})"
        render_metric_card("Top Section", top_section, "Highest final manpower")

    chart_col_1, chart_col_2 = st.columns([1.4, 1])

    with chart_col_1:
        # st.markdown('<div class="panel-card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">Section-wise final manpower</div>', unsafe_allow_html=True)
        st.markdown(
            '<div class="section-subtitle">Quick comparison of final manpower across the selected sections.</div>',
            unsafe_allow_html=True,
        )

        if filtered_summary_df.empty:
            st.info("No sections selected. Please choose one or more sections.")
        else:
            bar_chart = (
                alt.Chart(filtered_summary_df)
                .mark_bar(cornerRadiusTopLeft=8, cornerRadiusTopRight=8)
                .encode(
                    x=alt.X("Section:N", sort="-y", title="Section"),
                    y=alt.Y("BE_Final_Manpower:Q", title="Final manpower"),
                    tooltip=["Location", "Business", "Section", "BE_Final_Manpower"],
                )
                .properties(height=360)
            )
            st.altair_chart(bar_chart, width="stretch")

        st.markdown("</div>", unsafe_allow_html=True)

    with chart_col_2:
        # st.markdown('<div class="panel-card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">Shift mix</div>', unsafe_allow_html=True)
        st.markdown(
            '<div class="section-subtitle">Plant-level shift allocation from the current full table.</div>',
            unsafe_allow_html=True,
        )

        shift_summary = pd.DataFrame(
            {
                "Shift": ["General", "A", "B", "C"],
                "Manpower": [
                    pd.to_numeric(full_spinning_df["General_Shift"], errors="coerce").fillna(0).sum(),
                    pd.to_numeric(full_spinning_df["Shift_A"], errors="coerce").fillna(0).sum(),
                    pd.to_numeric(full_spinning_df["Shift_B"], errors="coerce").fillna(0).sum(),
                    pd.to_numeric(full_spinning_df["Shift_C"], errors="coerce").fillna(0).sum(),
                ],
            }
        )

        if shift_summary["Manpower"].sum() <= 0:
            st.info("Shift data is not available.")
        else:
            shift_chart = (
                alt.Chart(shift_summary)
                .mark_arc(innerRadius=58, outerRadius=108)
                .encode(
                    theta=alt.Theta("Manpower:Q"),
                    color=alt.Color("Shift:N", legend=alt.Legend(orient="bottom")),
                    tooltip=["Shift", "Manpower"],
                )
                .properties(height=350)
            )
            st.altair_chart(shift_chart, width="stretch")

        st.markdown("</div>", unsafe_allow_html=True)

    # st.markdown('<div class="panel-card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Summary table</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="section-subtitle">Section-wise final manpower ready for leadership review and download.</div>',
        unsafe_allow_html=True,
    )

    if filtered_summary_df.empty:
        st.info("No rows to display for the current filter selection.")
    else:
        st.dataframe(
            filtered_summary_df,
            width="stretch",
            hide_index=True,
            column_config={
                "BE_Final_Manpower": st.column_config.NumberColumn(
                    "BE_Final_Manpower",
                    format="%d",
                )
            },
        )

    current_lower_for_download = full_spinning_df.loc[
        full_spinning_df["Section"].astype(str).str.upper() == "TFO",
        DISPLAY_COLUMNS,
    ].copy()

    st.download_button(
        "Download final Excel",
        data=create_download_workbook(
            summary_df=build_summary_table(full_spinning_df),
            full_spinning_df=full_spinning_df,
            upper_tfo_df=current_upper_df,
            lower_tfo_df=current_lower_for_download,
        ),
        file_name="Spinning_Manpower_Final.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown("</div>", unsafe_allow_html=True)

    # ✅ ADD THIS BLOCK
    deviation_df = pd.read_excel("data/deviation.xlsx")

    st.markdown('<div class="section-title">Deviation Table</div>', unsafe_allow_html=True)

    st.dataframe(deviation_df, width="stretch", hide_index=True)


# -----------------
# VAPI
# -----------------
with vapi_tab:
    render_location_tab("Vapi", full_spinning_df, source_spinning_df)

# -----------------
# ANJAR
# -----------------
with anjar_tab:
    render_location_tab("Anjar", full_spinning_df, source_spinning_df)

# -----------------
# WFL
# -----------------
with wfl_tab:
    render_location_tab("Hyderabad", full_spinning_df, source_spinning_df)


# with spinning_tab:

#     sub_tab1, sub_tab2 = st.tabs(["Main", "TFO"])

#     # --- Main Spinning ---
#     with sub_tab1:
#         st.markdown('<div class="section-title">Entire Spinning Table</div>', unsafe_allow_html=True)

#         spinning_df = full_spinning_df[
#             full_spinning_df["Business"].str.upper() == "SPINNING"
#         ]

#         st.dataframe(spinning_df[DISPLAY_COLUMNS], width="stretch", hide_index=True)

#     with sub_tab2:
#     # st.markdown('<div class="panel-card">', unsafe_allow_html=True)
#         st.markdown('<div class="section-title">TFO planning and manpower engine</div>', unsafe_allow_html=True)
#         st.markdown(
#             '<div class="section-subtitle">Editable TFO production inputs with automatic roll-through into the entire spinning table and final summary.</div>',
#             unsafe_allow_html=True,
#         )

#         input_columns = [
#             "Count",
#             "Customer",
#             "Count2",
#             "Speed",
#             "TPI",
#             "Utilization",
#             "Efficiency",
#             "Production Required / day Kgs",
#             "TFO divisor",
#             "mpm",
#             "Eff",
#             "Machine divisor",
#         ]

#         edited_tfo_input_df = st.data_editor(
#             st.session_state.tfo_input_df[input_columns],
#             width="stretch",
#             num_rows="dynamic",
#             key="tfo_editor",
#             hide_index=True,
#             column_config={
#                 "Count": st.column_config.TextColumn("Count"),
#                 "Customer": st.column_config.TextColumn("Customer"),
#                 "Count2": st.column_config.NumberColumn("Count2", format="%.2f"),
#                 "Speed": st.column_config.NumberColumn("Speed", format="%.2f"),
#                 "TPI": st.column_config.NumberColumn("TPI", format="%.2f"),
#                 "Utilization": st.column_config.NumberColumn("Utilization", format="%.2f"),
#                 "Efficiency": st.column_config.NumberColumn("Efficiency", format="%.2f"),
#                 "Production Required / day Kgs": st.column_config.NumberColumn("Production Required / day Kgs", format="%.2f"),
#                 "TFO divisor": st.column_config.NumberColumn("TFO divisor", format="%.2f"),
#                 "mpm": st.column_config.NumberColumn("mpm", format="%.2f"),
#                 "Eff": st.column_config.NumberColumn("Eff", format="%.2f"),
#                 "Machine divisor": st.column_config.NumberColumn("Machine divisor", format="%.2f"),
#             },
#         )

#         for col in [
#             "Count2",
#             "Speed",
#             "TPI",
#             "Utilization",
#             "Efficiency",
#             "Production Required / day Kgs",
#             "TFO divisor",
#             "mpm",
#             "Eff",
#             "Machine divisor",
#         ]:
#             edited_tfo_input_df[col] = pd.to_numeric(
#                 edited_tfo_input_df[col],
#                 errors="coerce",
#             ).fillna(0).round(2)

#         st.session_state.tfo_input_df = edited_tfo_input_df.copy()

#         current_upper_df = calculate_upper_tfo_metrics(edited_tfo_input_df)
#         current_upper_total_df = build_upper_total_row(current_upper_df)
#         current_upper_final_df = pd.concat(
#             [current_upper_df.drop(columns=["Upper_Row_No"], errors="ignore"), current_upper_total_df],
#             ignore_index=True,
#         )

#         updated_full_df, current_lower_df, current_driver_values = rebuild_full_spinning_with_tfo(
#             source_df=source_spinning_df,
#             base_full_df=st.session_state.full_spinning_df,
#             upper_tfo_df=current_upper_df,
#         )

#         st.session_state.full_spinning_df = updated_full_df
#         full_spinning_df = updated_full_df.copy()

#         tfo_metric_1, tfo_metric_2, tfo_metric_3, tfo_metric_4 = st.columns(4)

#         with tfo_metric_1:
#             render_metric_card(
#                 "Total No. of Drums",
#                 f"{current_driver_values['sum_no_of_drums_total']:.2f}",
#                 "Calculated from inputs",
#             )

#         with tfo_metric_2:
#             render_metric_card(
#                 "TFO Required / Shift",
#                 f"{current_driver_values['sum_tfo_required_shift_total']:.2f}",
#                 "Based on divisor logic",
#             )

#         with tfo_metric_3:
#             render_metric_card(
#                 "Drums for 4/2/6 K",
#                 f"{current_driver_values['no_of_drums_426k']:.2f}",
#                 "Used in Jumbo Assembly",
#             )

#         with tfo_metric_4:
#             render_metric_card(
#                 "Lower TFO Final Manpower",
#                 f"{int(round(pd.to_numeric(current_lower_df['BE_Final_Manpower'], errors='coerce').fillna(0).sum())):,}",
#                 "Current TFO rows",
#             )

#         st.markdown("#### Upper TFO Production Table")

#         upper_display_columns = [
#             "Count",
#             "Customer",
#             "Count2",
#             "Speed",
#             "TPI",
#             "Utilization",
#             "Efficiency",
#             "Production per Drum/day",
#             "Production Required / day Kgs",
#             "Production Required / Month Kgs",
#             "No. of Drums Required",
#             "No. of TFO Required / shift",
#             "mpm",
#             "Eff",
#             "kgs/drum/day",
#             "No. of Drums",
#             "no. of machines",
#         ]

#         st.dataframe(
#             current_upper_final_df[upper_display_columns],
#             width="stretch",
#             hide_index=True,
#         )

#         st.markdown("#### Lower TFO Manpower Table")
#         st.dataframe(
#             current_lower_df,
#             width="stretch",
#             hide_index=True,
#             column_config={
#                 "BE_Scientific_Manpower": st.column_config.NumberColumn("BE_Scientific_Manpower", format="%.2f"),
#                 "BE_Final_Manpower": st.column_config.NumberColumn("BE_Final_Manpower", format="%.2f"),
#             },
#         )

#         with st.expander("Formula logic"):
#             st.markdown(
#                 f"""
#                 **Upper table**
#                 - Production Required / Month Kgs = Production Required / day Kgs × 30
#                 - Production per Drum/day = `((Speed × 60 × 8 × Efficiency × 2) / (TPI × 36 × 840 × Count2 × 2.202)) × 3`
#                 - No. of Drums Required = Production Required / day Kgs / Production per Drum/day
#                 - No. of TFO Required / shift = No. of Drums Required / TFO divisor
#                 - kgs/drum/day = `((mpm × Eff × 8 × 60 × 1.09) / (Count2 × 840 × 2.202)) × 3`
#                 - No. of Drums = Production Required / day Kgs / kgs/drum/day
#                 - no. of machines = No. of Drums / Machine divisor

#                 **Lower table**
#                 - Assembly winding = `ROUNDUP((SUM(T2:T18)/36)*3,0)`
#                 - Jumbo Assembly Winding = `ROUND(ROUNDUP(T19,0)/16,0)*2`
#                 - TFO Operator = `ROUND(SUM(N2:N18)/6,0)*3`
#                 - TFO Operator (Doffer) = `ROUNDUP(SUM(N2:N18)/4,0)*3`

#                 **Current driver values**
#                 - Total No. of Drums = {current_driver_values['sum_no_of_drums_total']:.2f}
#                 - Total No. of TFO Required / shift = {current_driver_values['sum_tfo_required_shift_total']:.2f}
#                 - No. of Drums for 4/2/6 K = {current_driver_values['no_of_drums_426k']:.2f}
#                 """
#             )

#         action_col_1, action_col_2 = st.columns(2)

#         with action_col_1:
#             if st.button("Reset TFO Table"):
#                 st.session_state.tfo_input_df = get_initial_tfo_data()
#                 reset_upper_df = calculate_upper_tfo_metrics(st.session_state.tfo_input_df)
#                 reset_full_df, _, _ = rebuild_full_spinning_with_tfo(
#                     source_df=source_spinning_df,
#                     base_full_df=st.session_state.full_spinning_df,
#                     upper_tfo_df=reset_upper_df,
#                 )
#                 st.session_state.full_spinning_df = reset_full_df
#                 st.session_state.full_spinning_editor_version += 1
#                 st.rerun()

#         with action_col_2:
#             if st.button("Reset Full Spinning Table from Source"):
#                 fresh_upper_df = calculate_upper_tfo_metrics(st.session_state.tfo_input_df)
#                 fresh_lower_df, _ = calculate_lower_tfo_manpower(fresh_upper_df)
#                 st.session_state.full_spinning_df = get_initial_full_spinning_df(source_spinning_df, fresh_lower_df)
#                 st.session_state.full_spinning_df = recalculate_scientific_manpower(
#                     st.session_state.full_spinning_df,
#                     fresh_upper_df,
#                 )
#                 tfo_mask = st.session_state.full_spinning_df["Section"].astype(str).str.upper() == "TFO"
#                 st.session_state.full_spinning_df.loc[tfo_mask, "BE_Final_Manpower"] = (
#                     st.session_state.full_spinning_df.loc[tfo_mask, "BE_Scientific_Manpower"]
#                 )
#                 st.session_state.full_spinning_editor_version += 1
#                 st.rerun()

#         st.markdown("</div>", unsafe_allow_html=True)


#         # --- STEP 4: Recalculate scientific manpower ---
#         # --- STEP 4: Recalculate scientific manpower ---
#         current_upper_df = calculate_upper_tfo_metrics(st.session_state.tfo_input_df)

#         st.session_state.full_spinning_df = recalculate_scientific_manpower(
#             st.session_state.full_spinning_df,
#             current_upper_df
#         )


# with wtt_tab:

#     wtt_prod = pd.ExcelFile("data/WTT.xlsx")
#     tabs = st.tabs([
#         "Manning",
#         "Weaving Back-Up",
#         "Stenter",
#         "Size Wise Details",
#         "TT Cut&Sew LC",
#         "TT Cut&Sew LH",
#         "DTA",
#         "JUKI"
#     ])
#     with tabs[0]:
#         st.markdown('<div class="section-title">Entire WTT Table</div>', unsafe_allow_html=True)

#         wtt_df = full_spinning_df[
#             full_spinning_df["Business"].str.upper() == "WTT"
#         ]
#         st.dataframe(wtt_df[DISPLAY_COLUMNS], width="stretch", hide_index=True)

#     with tabs[1]:
#         df = pd.read_excel(wtt_prod, sheet_name="Weaving Back-Up")
#         st.dataframe(df, width="stretch", hide_index=True)

#     with tabs[2]:
#         df = pd.read_excel(wtt_prod, sheet_name="Stenter")

#         st.markdown("### Dynamic Shift Planning Based on MT / Day")
#         st.dataframe(df.iloc[1:7, 0:2], width="stretch", hide_index=True)

#         st.markdown("#### Per Shift")
#         # df2 = df.iloc[8:13,0:4].rename(columns=df.iloc[8,0:4])
#         st.dataframe(df.iloc[9:13,0:4].rename(columns=df.iloc[8,0:4]), width="stretch", hide_index=True)

#     with tabs[3]:
#         df1 = pd.read_excel(wtt_prod, sheet_name="Size_wise_details")
#         df2 = pd.read_excel(wtt_prod, sheet_name="Size_wise_details_summary")

#         st.markdown("### Size Wise Details")
#         st.dataframe(df1, width="stretch", hide_index=True)

#         st.markdown("### Summary")
#         st.dataframe(df2.iloc[0:10, 0:6], width="stretch", hide_index=True)
        
#     with tabs[4]:
#         df = pd.read_excel(wtt_prod, sheet_name="TT_Cut&Sew_LC")

#         st.markdown("### Table 1")
#         st.dataframe(df.iloc[0:16, 0:10], width="stretch", hide_index=True)

#         st.markdown("### Table 2")
#         st.dataframe(df.iloc[16:18, 8:10], width="stretch", hide_index=True)

#     with tabs[5]:
#         df = pd.read_excel(wtt_prod, sheet_name="TT_Cut&Sew_LH")

#         st.markdown("### Table 1")
#         st.dataframe(df.iloc[0:17, 0:10], width="stretch", hide_index=True)

#         st.markdown("### Table 2")
#         st.dataframe(df.iloc[17:20, 8:10], width="stretch", hide_index=True)

#     with tabs[6]:
#         df = pd.read_excel(wtt_prod, sheet_name="DTA")
#         st.dataframe(df, width="stretch", hide_index=True)

#     with tabs[7]:
#         df = pd.read_excel(wtt_prod, sheet_name="JUKI")

#         st.markdown("### Table 1")
#         st.dataframe(df.iloc[0:14, 0:13], width="stretch", hide_index=True)

#         st.markdown("### Table 2")
#         st.dataframe(df.iloc[0:9, 13:15], width="stretch", hide_index=True)

    

# with rugs_tab:

#     rugs_sub1 = st.tabs(["Main"])[0]

#     with rugs_sub1:
#         st.markdown('<div class="section-title">Entire Rugs Table</div>', unsafe_allow_html=True)

#         rugs_df = full_spinning_df[
#             full_spinning_df["Business"].str.upper() == "RUGS"
#         ]

#         st.dataframe(rugs_df[DISPLAY_COLUMNS], width="stretch", hide_index=True)